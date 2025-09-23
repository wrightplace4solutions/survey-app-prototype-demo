import os
import pandas as pd
import smtplib
from email.mime.text import MIMEText
from openpyxl import load_workbook

def export_to_excel(data, filename="survey_results.xlsx", sheet_name="responses"):
    """
    Append a single submission (dict) to an Excel file.
    - Creates the file/sheet with headers on first run.
    - Appends subsequent rows on later runs.
    """
    df = pd.DataFrame([data])

    if os.path.exists(filename):
        # Append to existing workbook/sheet
        with pd.ExcelWriter(
            filename, engine="openpyxl", mode="a", if_sheet_exists="overlay"
        ) as writer:
            book = writer.book
            if sheet_name in book.sheetnames:
                ws = book[sheet_name]
                startrow = ws.max_row
            else:
                # New sheet in existing workbook
                startrow = 0
            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=(startrow == 0),
                startrow=startrow,
            )
    else:
        # Create new workbook with the sheet and headers
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    return filename

def send_email(subject, body, to_emails):
    msg = MIMEText(body)
    msg["Subject"] = subject
    msg["From"] = "noreply@training-survey-app.com"
    msg["To"] = ", ".join(to_emails)

    try:
        with smtplib.SMTP("localhost") as server:
            server.send_message(msg)
    except Exception as e:
        print("Email send failed:", e)
